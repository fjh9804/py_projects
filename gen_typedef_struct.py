import re
from openpyxl import load_workbook

wb = load_workbook("input.xlsx")
ws = wb.active
out_ws = wb.create_sheet("Parsed")

row_out = 1

for row in ws.iter_rows(min_col=2, max_col=2):
    cell = row[0].value
    if not cell:
        continue

    text = cell.strip()

    # 提取结构体名
    struct_name_match = re.search(r"}\s*(\w+)\s*;", text)
    if not struct_name_match:
        continue
    struct_name = struct_name_match.group(1)

    # 提取结构体体内内容
    body_match = re.search(r"{([\s\S]*?)}", text)
    if not body_match:
        continue

    body = body_match.group(1)

    # 提取成员变量（简单规则）
    members = re.findall(r"\b\w+\s+(\w+)\s*;", body)

    # 写结构体名（第一行）
    out_ws.cell(row=row_out, column=1, value=struct_name)

    first = True
    for m in members:
        if not first:
            row_out += 1
        out_ws.cell(row=row_out, column=2, value=m)
        first = False

    # 空行分隔
    row_out += 2

wb.save("output.xlsx")